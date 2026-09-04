from datetime import date, timedelta

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import load_workbook

from apps.amm.models import MarketingAuthorization
from tests.conftest import TODAY

pytestmark = pytest.mark.django_db


def test_analytics_africa(hq_client, country_client, make_amm):
    make_amm(country="SN")  # valid, OK
    make_amm(country="SN", start=TODAY - timedelta(days=365 * 5 - 100))  # valid, expiring < 6m
    make_amm(country="SN", start=date(2015, 1, 1))  # expired
    make_amm(country="SN", original_start_date=None, dossier_state="INCONNU")  # undetermined
    make_amm(country="CI", dossier_state="INCOMPLET")
    body = hq_client.get("/api/v1/analytics/africa").json()
    rows = {row["country_iso2"]: row for row in body["rows"]}
    assert rows["SN"]["total"] == 4 and rows["SN"]["valid"] == 2 and rows["SN"]["expired"] == 1
    assert (
        rows["SN"]["undetermined"] == 1
        and rows["SN"]["expiring_6m"] == 1
        and rows["SN"]["expiring_12m"] == 1
    )
    assert rows["SN"]["pct_valid"] == 50.0 and rows["SN"]["pct_complete"] == 75.0
    assert body["total"]["total"] == 5 and body["total"]["valid"] == 3
    assert rows["ML"]["total"] == 0
    scoped = country_client.get("/api/v1/analytics/africa").json()
    assert {r["country_iso2"] for r in scoped["rows"]} == {"SN", "ML"}
    assert scoped["total"]["total"] == 4


def test_analytics_country_and_coverage(hq_client, country_client, make_amm, product):
    make_amm(country="SN", product_obj=product, start=TODAY - timedelta(days=365 * 5 - 40))
    make_amm(country="CI", product_obj=product, start=date(2015, 1, 1))
    body = hq_client.get("/api/v1/analytics/country/SN").json()
    assert body["country"]["iso2"] == "SN"
    assert len(body["pipeline"]) == 24 and body["pipeline"][0]["month"] == TODAY.strftime("%Y-%m")
    assert sum(p["count"] for p in body["pipeline"]) == 1
    assert body["by_range_status"] == [{"range": "GENERALE", "status": "VALIDE", "count": 1}]
    assert body["priorities"][0]["urgency"] == "CRITIQUE"
    assert country_client.get("/api/v1/analytics/country/CI").status_code == 403

    coverage = hq_client.get(f"/api/v1/analytics/product/{product.pk}/coverage").json()
    by_country = {c["country_iso2"]: c for c in coverage}
    assert by_country["SN"]["status"] == "VALIDE" and by_country["CI"]["status"] == "EXPIRE"
    assert by_country["ML"]["status"] is None and len(coverage) == 4


def test_export_xlsx_and_csv(hq_client, make_amm, make_renewal):
    amm = make_amm(country="SN", start=date(2019, 1, 1))
    make_renewal(amm, "OBTENU", number="R-1", start_date=date(2024, 1, 1))
    make_amm(country="CI")
    response = hq_client.get("/api/v1/analytics/export?format=xlsx&country=SN")
    assert response.status_code == 200
    sheet = load_workbook(filename=__import__("io").BytesIO(response.content)).active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == (
        "PAYS",
        "GAMME",
        "NOM",
        "DATE DEBUT",
        "N° AMM",
        "DATE FIN",
        "RENOUV. DATE DEBUT",
        "RENOUV. N° AMM",
        "RENOUV. DATE FIN",
        "STATUT",
        "ETAT DOSSIER",
        "URGENCE",
        "DEADLINE DEPOT",
    )
    assert (
        len(rows) == 2
        and rows[1][0] == "Sénégal"
        and rows[1][7] == "R-1"
        and rows[1][9] == "VALIDE"
    )
    csv_response = hq_client.get("/api/v1/analytics/export?format=csv")
    assert csv_response.status_code == 200
    lines = csv_response.content.decode("utf-8").strip().splitlines()
    assert lines[0].startswith("PAYS;GAMME;NOM") and len(lines) == 3


def test_import_endpoint(hq_client, country_client, workbook_bytes, ranges):
    upload = SimpleUploadedFile("classeur.xlsx", workbook_bytes)
    assert (
        country_client.post("/api/v1/imports", {"file": upload}, format="multipart").status_code
        == 403
    )
    upload = SimpleUploadedFile("classeur.xlsx", workbook_bytes)
    response = hq_client.post(
        "/api/v1/imports", {"file": upload, "today": TODAY.isoformat()}, format="multipart"
    )
    assert response.status_code == 202, response.json()
    batch = response.json()
    assert batch["status"] == "DONE"  # eager Celery in tests
    assert batch["summary"]["totals"]["created"] == 7
    assert MarketingAuthorization.objects.count() == 7
    detail = hq_client.get(f"/api/v1/imports/{batch['id']}").json()
    assert detail["summary"]["sheets"]["MALI"]["country"] == "ML"
    errors = hq_client.get(f"/api/v1/imports/{batch['id']}/rows?outcome=ERROR").json()
    assert errors["count"] == 1 and "date illisible" in errors["results"][0]["message"]
    assert hq_client.get(f"/api/v1/imports/{batch['id']}/rows").json()["count"] == 7


def test_users_crud_permissions(ceo_client, hq_client, country_client, users, countries):
    assert country_client.get("/api/v1/users").status_code == 200
    assert country_client.get("/api/v1/users").json()["count"] == 0
    assert (
        country_client.post(
            "/api/v1/users", {"email": "x@test.local", "role": "COUNTRY_REGULATORY"}
        ).status_code
        == 403
    )

    created = hq_client.post(
        "/api/v1/users",
        {
            "email": "Mali@test.local",
            "role": "COUNTRY_REGULATORY",
            "countries": ["ML"],
            "password": "Passw0rd!",
        },
    )
    assert created.status_code == 201, created.json()
    assert created.json()["email"] == "mali@test.local" and created.json()["countries"] == ["ML"]
    forbidden = hq_client.post("/api/v1/users", {"email": "boss@test.local", "role": "CEO_ADMIN"})
    assert forbidden.status_code == 403
    assert (
        hq_client.patch(f"/api/v1/users/{users['ceo'].pk}", {"first_name": "X"}).status_code == 403
    )
    assert (
        hq_client.patch(
            f"/api/v1/users/{created.json()['id']}", {"countries": ["ML", "SN"]}
        ).status_code
        == 200
    )
    assert (
        hq_client.patch(
            f"/api/v1/users/{created.json()['id']}", {"role": "HQ_REGULATORY"}
        ).status_code
        == 403
    )

    promoted = ceo_client.post(
        "/api/v1/users",
        {"email": "hq2@test.local", "role": "HQ_REGULATORY", "password": "Passw0rd!"},
    )
    assert promoted.status_code == 201
    assert ceo_client.delete(f"/api/v1/users/{promoted.json()['id']}").status_code == 204
    assert ceo_client.get("/api/v1/users?role=COUNTRY_REGULATORY").json()["count"] == 2
    weak = ceo_client.post(
        "/api/v1/users", {"email": "w@test.local", "role": "COUNTRY_REGULATORY", "password": "123"}
    )
    assert weak.status_code == 400


def test_pagination_page_size(hq_client, make_amm):
    for _ in range(3):
        make_amm()
    body = hq_client.get("/api/v1/amms?page_size=2").json()
    assert len(body["results"]) == 2 and body["next"] is not None and body["count"] == 3
