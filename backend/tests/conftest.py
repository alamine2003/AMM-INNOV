"""Shared fixtures: fixed reference date, referentials, users, JWT clients, factories."""

import io
from datetime import date, timedelta

import pytest
from openpyxl import Workbook
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import RefreshToken

from apps.accounts.models import User
from apps.alerts.defaults import seed_default_rules
from apps.alerts.models import AlertRule
from apps.amm.models import MarketingAuthorization, Renewal
from apps.catalog.models import Country, Product, ProductRange
from apps.core.dates import override_today

TODAY = date(2026, 9, 4)

MINIMAL_PDF = (
    b"%PDF-1.4\n1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n"
    b"2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n"
    b"3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 200 200]>>endobj\n"
    b"xref\n0 4\n0000000000 65535 f \ntrailer<</Size 4/Root 1 0 R>>\nstartxref\n0\n%%EOF\n"
)


@pytest.fixture(autouse=True)
def fixed_today():
    with override_today(TODAY):
        yield TODAY


@pytest.fixture
def countries(db):
    return {
        "SN": Country.objects.create(iso2="SN", name="Sénégal", authority="DPM"),
        "ML": Country.objects.create(iso2="ML", name="Mali"),
        "CI": Country.objects.create(iso2="CI", name="Côte d'Ivoire"),
        "GN": Country.objects.create(iso2="GN", name="Guinée", validity_years=3),
    }


@pytest.fixture
def ranges(db):
    return {
        code: ProductRange.objects.create(code=code, label=label)
        for code, label in ProductRange.Code.choices
    }


@pytest.fixture
def product(ranges):
    return Product.objects.create(name="ARTEGEN 120MG PDRE SOL INJ", range=ranges["GENERALE"])


@pytest.fixture
def rules(db):
    seed_default_rules(AlertRule)
    return {rule.code: rule for rule in AlertRule.objects.filter(country=None)}


@pytest.fixture
def users(countries):
    ceo = User.objects.create_user("ceo@test.local", "Passw0rd!", role=User.Role.CEO_ADMIN)
    hq = User.objects.create_user("hq@test.local", "Passw0rd!", role=User.Role.HQ_REGULATORY)
    country = User.objects.create_user(
        "sn@test.local", "Passw0rd!", role=User.Role.COUNTRY_REGULATORY, first_name="Fatou"
    )
    country.countries.set([countries["SN"], countries["ML"]])
    return {"ceo": ceo, "hq": hq, "country": country}


def client_for(user: User) -> APIClient:
    client = APIClient()
    token = RefreshToken.for_user(user).access_token
    client.credentials(HTTP_AUTHORIZATION=f"Bearer {token}")
    return client


@pytest.fixture
def ceo_client(users):
    return client_for(users["ceo"])


@pytest.fixture
def hq_client(users):
    return client_for(users["hq"])


@pytest.fixture
def country_client(users):
    return client_for(users["country"])


@pytest.fixture
def anon_client():
    return APIClient()


@pytest.fixture
def make_amm(countries, product, ranges):
    counter = {"n": 0}

    def _make(country="SN", start=None, product_obj=None, **kwargs):
        counter["n"] += 1
        if product_obj is None:
            product_obj = Product.objects.create(
                name=f"PRODUIT TEST {counter['n']}", range=ranges["GENERALE"]
            )
        if "original_start_date" in kwargs:
            start = kwargs.pop("original_start_date")
        elif start is None:
            start = TODAY - timedelta(days=365)
        kwargs.setdefault("original_number", f"N-{counter['n']:04d}")
        kwargs.setdefault("dossier_state", "COMPLET")
        return MarketingAuthorization.objects.create(
            product=product_obj,
            country=countries[country] if isinstance(country, str) else country,
            original_start_date=start,
            **kwargs,
        )

    return _make


@pytest.fixture
def make_renewal():
    def _make(amm, status="PLANIFIE", **kwargs):
        return Renewal.objects.create(amm=amm, workflow_status=status, **kwargs)

    return _make


def build_workbook() -> bytes:
    """Synthetic workbook: DASHBOARD + 2 normalized sheets + 1 old-format sheet."""
    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "DASHBOARD"
    dashboard["B2"] = "TABLEAU DE BORD AMM - AFRIQUE"

    header = [
        "NBR",
        "NBR",
        "GAMME",
        "NOM",
        "DATE DEBUT",
        "N° AMM",
        "DATE FIN",
        "DATE DEBUT",
        "N° AMM",
        "DATE FIN",
        "STATUT",
        "ETAT DOSSIER",
    ]
    senegal = workbook.create_sheet("SENEGAL ")
    senegal.append([None, None, None, None, "AMM D'ORIGINE", None, None, "DERNIER RENOUVELLEMENT"])
    senegal.append(header)
    senegal.append(
        [
            1,
            1,
            "GAMME GENERALE",
            "ARTEGEN 120MG PDRE SOL INJ",
            date(2025, 4, 28),
            "2025-0001040",
            date(2030, 4, 28),
            None,
            None,
            None,
            "VALIDE",
            "Dossier complet",
        ]
    )
    senegal.append(
        [
            2,
            2,
            "GAMME GENERAL",
            "ALFA GH SR 10MG CPR B/30",
            date(2020, 1, 13),
            5859.0,
            date(2025, 1, 13),
            date(2025, 1, 13),
            5859.0,
            date(2030, 1, 13),
            "VALIDE",
            "Dossier complet",
        ]
    )
    senegal.append(
        [
            3,
            3,
            "GAME CARDIO",
            "AMLO VH 10MG CPR B/30",
            date(2020, 6, 1),
            "X-1",
            date(2025, 6, 1),
            "IN PROCESS",
            "X-1",
            None,
            "IN PROCESS",
            "Dossier incomplet",
        ]
    )
    senegal.append(
        [
            4,
            4,
            "GAMME BIEN ETRE",
            "GENSIL HUILE F60ML",
            "DATE ILLISIBLE",
            None,
            None,
            None,
            None,
            None,
            "INDETERMINE",
            None,
        ]
    )
    senegal.append(
        [
            5,
            5,
            "GAMME GENERALE",
            "KETOPROFEN GH 100MG CPR B30",
            date(2019, 3, 1),
            "K-1",
            date(2024, 3, 1),
            None,
            None,
            None,
            "VALIDE",
            "Dossier complet",
        ]
    )
    senegal.append([None, None, None, None])

    mali = workbook.create_sheet("MALI")
    mali.append([None])
    mali.append(["NBR", "NBR/PDT"] + header[2:])
    mali.append(
        [
            1,
            1,
            "GAMME GENERALE",
            "ARTEGEN 120MG PDRE SOL INJ",
            date(2024, 2, 21),
            "245 018",
            date(2029, 2, 21),
            None,
            None,
            None,
            "VALIDE",
            "Dossier complet",
        ]
    )
    mali.append(
        [
            2,
            2,
            "GAMME CARDIO",
            "RAMIPRIL GH 5MG CPR B/30",
            date(2018, 5, 5),
            "R-5",
            date(2023, 5, 5),
            None,
            None,
            None,
            "EXPIRE",
            "Dossier incomplet",
        ]
    )

    old = workbook.create_sheet("BENIN")
    old.append(["GAMME GENERALE"])
    old.append([None, None, "AMM D'ORIGINE"])
    old.append(["NBR", "NOM", "DATE DEBUT", "N° AMM", "DATE FIN"])
    old.append([1, "VIEUX FORMAT", date(2020, 1, 1), "OLD-1", date(2025, 1, 1)])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


@pytest.fixture
def workbook_bytes():
    return build_workbook()


@pytest.fixture
def workbook_path(tmp_path, workbook_bytes):
    path = tmp_path / "classeur.xlsx"
    path.write_bytes(workbook_bytes)
    return path
