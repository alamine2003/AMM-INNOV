"""Parser of the "Dashboard AMM Afrique" workbook (section 5.9 of the architecture).

Only sheets whose row 2 matches the normalized 12-column header are retained; data starts
on row 3 and stops at the first empty product name (column D). Parsing is pure: nothing is
written to the database here.
"""

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from apps.catalog.normalize import normalize_product_name, normalize_range

SHEET_COUNTRIES: dict[str, tuple[str, str]] = {
    "BENIN": ("BJ", "Bénin"),
    "BURKINA": ("BF", "Burkina Faso"),
    "CAMEROUN": ("CM", "Cameroun"),
    "CDI": ("CI", "Côte d'Ivoire"),
    "CONGO": ("CG", "Congo"),
    "DJIBOUTI": ("DJ", "Djibouti"),
    "GABON": ("GA", "Gabon"),
    "GAMBIE": ("GM", "Gambie"),
    "GUINEE": ("GN", "Guinée"),
    "MALI": ("ML", "Mali"),
    "MADAGASCAR": ("MG", "Madagascar"),
    "NIGER": ("NE", "Niger"),
    "SENEGAL": ("SN", "Sénégal"),
    "TCHAD": ("TD", "Tchad"),
    "TOGO": ("TG", "Togo"),
}

HEADER = [
    "NBR",
    "NBR",
    "GAMME",
    "NOM",
    "DATE DEBUT",
    "N° AMM",
    "DATE FIN",
    "DATE DEBUT",
    "N° AMM",
    "DATE FIN",
    "STATUT",
    "ETAT DOSSIER",
]
COLUMNS = [
    "nbr",
    "nbr_pdt",
    "gamme",
    "nom",
    "date_debut",
    "num_amm",
    "date_fin",
    "renouv_date_debut",
    "renouv_num_amm",
    "renouv_date_fin",
    "statut",
    "etat_dossier",
]
DATE_FORMATS = ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%Y-%m-%d")
IN_PROCESS = "IN PROCESS"
STATUS_MAP = {
    "VALIDE": "VALIDE",
    "VALIDES": "VALIDE",
    "EXPIRE": "EXPIRE",
    "EXPIREE": "EXPIRE",
    "EXPIRÉ": "EXPIRE",
    "EXPIRÉE": "EXPIRE",
    "IN PROCESS": "IN_PROCESS",
    "IN_PROCESS": "IN_PROCESS",
    "EN COURS": "IN_PROCESS",
    "INDETERMINE": "INDETERMINE",
    "INDÉTERMINÉ": "INDETERMINE",
}


def _header_token(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value)).upper().replace("Nº", "N°").replace("NO", "N°")


def matches_normalized_header(cells: list[Any]) -> bool:
    """Row 2 must be the normalized header; A2/B2 tolerate `NBR`, `NBR/PDT` or empty."""
    tokens = [_header_token(c) for c in list(cells[:12]) + [None] * (12 - len(cells[:12]))]
    if tokens[0] != "NBR":
        return False
    if tokens[1] not in {"NBR", "NBR/PDT", ""}:
        return False
    return tokens[2:12] == [_header_token(h) for h in HEADER[2:]]


def parse_date(value: Any) -> tuple[date | None, str | None]:
    """Returns (date, error). Empty -> (None, None); unreadable -> (None, "date illisible ...")."""
    if value is None:
        return None, None
    if isinstance(value, datetime):
        return value.date(), None
    if isinstance(value, date):
        return value, None
    if isinstance(value, int | float) and not isinstance(value, bool):
        try:
            converted = from_excel(value)
            return (converted.date() if isinstance(converted, datetime) else converted), None
        except (ValueError, TypeError, OverflowError):
            return None, f"date illisible : {value!r}"
    text = str(value).strip()
    if not text:
        return None, None
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date(), None
        except ValueError:
            continue
    return None, f"date illisible : « {text} »"


def parse_number(value: Any) -> str:
    """AMM numbers are strings; floats coming from Excel lose their `.0`."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else repr(value)
    if isinstance(value, int):
        return str(value)
    return re.sub(r"\s+", " ", str(value)).strip()


def parse_status(value: Any) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip().upper()
    return STATUS_MAP.get(text, text or None)


def parse_dossier_state(value: Any) -> str:
    if value is None:
        return "INCONNU"
    text = str(value).strip().lower()
    if "incomplet" in text:
        return "INCOMPLET"
    if "complet" in text:
        return "COMPLET"
    return "INCONNU"


def _is_in_process(value: Any) -> bool:
    return isinstance(value, str) and re.sub(r"\s+", " ", value).strip().upper() == IN_PROCESS


def _json_safe(value: Any) -> Any:
    if isinstance(value, datetime | date):
        return value.isoformat()
    return value


@dataclass
class ParsedRow:
    row_number: int
    raw: dict[str, Any]
    raw_name: str
    product_name: str
    range_raw: str | None
    range_code: str | None
    original_number: str
    original_start_date: date | None
    original_end_date: date | None
    renewal_in_process: bool
    renewal_number: str
    renewal_start_date: date | None
    renewal_end_date: date | None
    excel_status: str | None
    dossier_state: str
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


@dataclass
class ParsedSheet:
    name: str
    country_iso2: str
    country_name: str
    rows: list[ParsedRow] = field(default_factory=list)


@dataclass
class ParsedWorkbook:
    sheets: list[ParsedSheet] = field(default_factory=list)
    ignored: list[dict[str, str]] = field(default_factory=list)


def parse_row(row_number: int, cells: list[Any]) -> ParsedRow:
    cells = list(cells[:12]) + [None] * (12 - len(cells[:12]))
    raw = {name: _json_safe(value) for name, value in zip(COLUMNS, cells, strict=True)}
    errors: list[str] = []
    warnings: list[str] = []

    raw_name = normalize_product_name(cells[3])
    range_code = normalize_range(cells[2])
    if range_code is None:
        warnings.append(f"gamme inconnue : « {cells[2]} »" if cells[2] else "gamme absente")

    original_start, err = parse_date(cells[4])
    if err:
        errors.append(f"colonne E (DATE DEBUT) — {err}")
    original_end, err = parse_date(cells[6])
    if err:
        errors.append(f"colonne G (DATE FIN) — {err}")

    in_process = _is_in_process(cells[7])
    renewal_start: date | None = None
    if not in_process:
        renewal_start, err = parse_date(cells[7])
        if err:
            errors.append(f"colonne H (RENOUV. DATE DEBUT) — {err}")
    renewal_end, err = parse_date(cells[9])
    if err:
        errors.append(f"colonne J (RENOUV. DATE FIN) — {err}")
    if not in_process and renewal_start is None and renewal_end is not None:
        warnings.append("renouvellement sans date de début : date de fin ignorée")

    return ParsedRow(
        row_number=row_number,
        raw=raw,
        raw_name=raw_name,
        product_name=raw_name,
        range_raw=str(cells[2]).strip() if cells[2] else None,
        range_code=range_code,
        original_number=parse_number(cells[5]),
        original_start_date=original_start,
        original_end_date=original_end,
        renewal_in_process=in_process,
        renewal_number=parse_number(cells[8]),
        renewal_start_date=renewal_start,
        renewal_end_date=renewal_end,
        excel_status=parse_status(cells[10]),
        dossier_state=parse_dossier_state(cells[11]),
        errors=errors,
        warnings=warnings,
    )


def parse_workbook(source) -> ParsedWorkbook:
    """`source` is a path or a file-like object."""
    workbook = load_workbook(source, data_only=True, read_only=True)
    result = ParsedWorkbook()
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(min_row=1, values_only=True)
            next(rows, None)  # row 1: block titles
            header = next(rows, None)
            if header is None or not matches_normalized_header(list(header)):
                result.ignored.append(
                    {
                        "sheet": sheet.title,
                        "reason": "en-tête non normalisé (ancien format ou onglet de synthèse)",
                    }
                )
                continue
            mapping = SHEET_COUNTRIES.get(sheet.title.strip().upper())
            if mapping is None:
                result.ignored.append(
                    {"sheet": sheet.title, "reason": "onglet inconnu : aucun pays associé"}
                )
                continue
            parsed = ParsedSheet(name=sheet.title, country_iso2=mapping[0], country_name=mapping[1])
            for row_number, cells in enumerate(rows, start=3):
                cells = list(cells)
                name = cells[3] if len(cells) > 3 else None
                if name is None or not str(name).strip():
                    break
                parsed.rows.append(parse_row(row_number, cells))
            result.sheets.append(parsed)
    finally:
        workbook.close()
    return result
