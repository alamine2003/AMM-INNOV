"""Excel and CSV exports of a filtered AMM queryset."""

import csv
import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .services import EXPORT_COLUMNS, export_rows


def _cell(value):
    if isinstance(value, date):
        return value
    return value if value is not None else ""


def build_xlsx(queryset) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "AMM"
    sheet.append(EXPORT_COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in export_rows(queryset):
        sheet.append([_cell(v) for v in row])
    for column in range(1, len(EXPORT_COLUMNS) + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 18
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, date):
                cell.number_format = "DD/MM/YYYY"
    sheet.freeze_panes = "A2"
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_csv(queryset) -> str:
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";")
    writer.writerow(EXPORT_COLUMNS)
    for row in export_rows(queryset):
        writer.writerow(
            [
                v.strftime("%d/%m/%Y") if isinstance(v, date) else ("" if v is None else v)
                for v in row
            ]
        )
    return buffer.getvalue()
